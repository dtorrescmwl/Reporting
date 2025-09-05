#!/usr/bin/env python3
"""
General CSV to XLSX Converter

A utility script to convert CSV files to XLSX format with proper formatting
and multiple sheet support. This is a general-purpose tool that can be used
across different systems.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import sys
import os

def convert_csv_to_xlsx(csv_path, xlsx_path, sheet_name="Data"):
    """Convert a single CSV file to XLSX format."""
    try:
        # Read CSV file
        df = pd.read_csv(csv_path)
        print(f"Read {len(df)} rows from {csv_path}")
        
        # Create workbook and add data
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add headers with formatting
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        
        # Add data
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column].width = adjusted_width
        
        # Save workbook
        wb.save(xlsx_path)
        print(f"Successfully converted to {xlsx_path}")
        return True
        
    except Exception as e:
        print(f"Error converting CSV to XLSX: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python csv_to_xlsx_converter.py input.csv output.xlsx")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_xlsx = sys.argv[2]
    
    if not os.path.exists(input_csv):
        print(f"Error: Input file {input_csv} not found")
        sys.exit(1)
    
    success = convert_csv_to_xlsx(input_csv, output_xlsx)
    sys.exit(0 if success else 1)