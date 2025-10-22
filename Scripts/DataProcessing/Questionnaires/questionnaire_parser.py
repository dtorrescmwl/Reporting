#!/usr/bin/env python3
"""
Questionnaire Parser
Parses questionnaire CSV data and organizes into separate Excel tabs by questionnaire type.
Usage: python3 questionnaire_parser.py <input_csv>
"""

import sys
import csv
import json
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import pytz


def extract_list_names(list_data):
    """Extract names from list of objects with 'name' field"""
    if not isinstance(list_data, list):
        return ""
    names = [item.get('name', '') for item in list_data if isinstance(item, dict) and 'name' in item]
    return ", ".join(names)


def consolidate_boolean_group(group_dict):
    """Consolidate boolean dictionary into comma-separated list of true values"""
    if not isinstance(group_dict, dict):
        return group_dict

    # Extract keys where value is True
    true_keys = []
    other_values = {}

    for key, value in group_dict.items():
        if isinstance(value, bool) and value:
            # Remove redundant prefix from key name
            clean_key = key.split('_')[-1] if '_' in key else key
            true_keys.append(clean_key)
        elif not isinstance(value, bool):
            # Keep non-boolean values (like descriptions)
            clean_key = key.split('_')[-1] if '_' in key else key
            other_values[clean_key] = value

    if not true_keys and not other_values:
        return ""

    result = ", ".join(true_keys) if true_keys else ""

    # Add descriptions or other values WITHOUT redundant key prefix
    for key, value in other_values.items():
        if result:
            result += f" ({key}: {value})"
        else:
            # Just return the value without the redundant key
            result = str(value)

    return result


def convert_to_eastern_time(utc_timestamp_str):
    """Convert UTC timestamp string to Eastern Time"""
    if not utc_timestamp_str:
        return ""

    try:
        # Parse ISO 8601 format
        utc_dt = datetime.fromisoformat(utc_timestamp_str.replace('Z', '+00:00'))

        # Convert to Eastern Time
        eastern = pytz.timezone('America/New_York')
        et_dt = utc_dt.astimezone(eastern)

        return et_dt.strftime('%Y-%m-%d %H:%M:%S %Z')
    except Exception as e:
        print(f"Warning: Could not convert timestamp {utc_timestamp_str}: {e}")
        return utc_timestamp_str


def should_exclude_column(column_name):
    """Check if column should be excluded (consent, complete, status, states)"""
    exclude_keywords = ['consent', 'complete', 'status', 'states']
    column_lower = column_name.lower()
    return any(keyword in column_lower for keyword in exclude_keywords)


def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary structure with smart consolidation"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k

        # Skip excluded fields
        if should_exclude_column(new_key):
            continue

        if isinstance(v, dict):
            # Check if this dict contains a 'list' key (medications, supplements, etc.)
            if 'list' in v and isinstance(v['list'], list):
                # Extract names from list
                items.append((new_key, extract_list_names(v['list'])))
            # Special handling for height/weight data
            elif 'heightUnit' in v or 'weightUnit' in v:
                # Keep height and weight fields separate
                for sub_key, sub_val in v.items():
                    items.append((f"{new_key}_{sub_key}", sub_val))
            # Check if this is a boolean group that should be consolidated
            elif all(isinstance(val, (bool, str, int, float)) for val in v.values()):
                consolidated = consolidate_boolean_group(v)
                items.append((new_key, consolidated))
            else:
                # Recursively flatten
                items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Convert lists to comma-separated if they're simple, otherwise JSON
            if all(isinstance(item, str) for item in v):
                items.append((new_key, ", ".join(v)))
            else:
                items.append((new_key, json.dumps(v)))
        else:
            items.append((new_key, v))
    return dict(items)


def extract_customer_info(customer_data):
    """Extract relevant customer information"""
    return {
        'user_id': customer_data.get('_id', ''),
        'gender': customer_data.get('gender', ''),
        'dob': customer_data.get('dob', '')
    }


def parse_wlus(entry):
    """Parse WLUS questionnaire data"""
    answers = entry.get('answers', {})
    flat_data = flatten_dict(answers)

    customer = entry.get('customer', {})
    first_name = customer.get('firstName', '')
    last_name = customer.get('lastName', '')
    full_name = f"{first_name} {last_name}".strip()

    # Add metadata
    result = {
        'questionnaire_id': entry.get('_id', ''),
        'user_id': customer.get('_id', ''),
        'name': full_name,
        'gender': customer.get('gender', ''),
        'dob': customer.get('dob', ''),
        'created_at': convert_to_eastern_time(entry.get('createdAt', '')),
        'updated_at': convert_to_eastern_time(entry.get('updatedAt', '')),
    }

    # Add all answer fields
    result.update(flat_data)

    return result


def parse_wlusmonthly(entry):
    """Parse WLUS Monthly questionnaire data"""
    answers = entry.get('answers', {})
    flat_data = flatten_dict(answers)

    customer = entry.get('customer', {})
    first_name = customer.get('firstName', '')
    last_name = customer.get('lastName', '')
    full_name = f"{first_name} {last_name}".strip()

    # Add metadata
    result = {
        'questionnaire_id': entry.get('_id', ''),
        'user_id': customer.get('_id', ''),
        'name': full_name,
        'gender': customer.get('gender', ''),
        'dob': customer.get('dob', ''),
        'created_at': convert_to_eastern_time(entry.get('createdAt', '')),
        'updated_at': convert_to_eastern_time(entry.get('updatedAt', '')),
    }

    # Add all answer fields
    result.update(flat_data)

    return result


def parse_facesheet(entry):
    """Parse Facesheet questionnaire data"""
    answers = entry.get('answers', {})
    flat_data = flatten_dict(answers)

    customer = entry.get('customer', {})
    first_name = customer.get('firstName', '')
    last_name = customer.get('lastName', '')
    full_name = f"{first_name} {last_name}".strip()

    # Add metadata
    result = {
        'questionnaire_id': entry.get('_id', ''),
        'user_id': customer.get('_id', ''),
        'name': full_name,
        'gender': customer.get('gender', ''),
        'dob': customer.get('dob', ''),
        'created_at': convert_to_eastern_time(entry.get('createdAt', '')),
        'updated_at': convert_to_eastern_time(entry.get('updatedAt', '')),
    }

    # Add all answer fields
    result.update(flat_data)

    return result


def parse_facesheet_ext(entry):
    """Parse Facesheet Extended questionnaire data"""
    answers = entry.get('answers', {})
    flat_data = flatten_dict(answers)

    customer = entry.get('customer', {})
    first_name = customer.get('firstName', '')
    last_name = customer.get('lastName', '')
    full_name = f"{first_name} {last_name}".strip()

    # Add metadata
    result = {
        'questionnaire_id': entry.get('_id', ''),
        'user_id': customer.get('_id', ''),
        'name': full_name,
        'gender': customer.get('gender', ''),
        'dob': customer.get('dob', ''),
        'created_at': convert_to_eastern_time(entry.get('createdAt', '')),
        'updated_at': convert_to_eastern_time(entry.get('updatedAt', '')),
    }

    # Add all answer fields
    result.update(flat_data)

    return result


def filter_duplicates_within_24h(entries):
    """Filter duplicate entries for same user within 24h, keeping only the newest"""
    if not entries:
        return entries

    # Group by user_id
    user_entries = {}
    for entry in entries:
        user_id = entry.get('user_id', '')
        if user_id not in user_entries:
            user_entries[user_id] = []
        user_entries[user_id].append(entry)

    filtered_entries = []
    duplicates_removed = 0

    for user_id, user_entry_list in user_entries.items():
        if len(user_entry_list) == 1:
            filtered_entries.append(user_entry_list[0])
            continue

        # Sort by created_at (newest first)
        sorted_entries = sorted(
            user_entry_list,
            key=lambda x: x.get('created_at', ''),
            reverse=True
        )

        # Keep track of entries to include
        included = [sorted_entries[0]]  # Always include the newest

        for entry in sorted_entries[1:]:
            # Check if this entry is within 24h of any included entry
            entry_time = entry.get('created_at', '')
            is_duplicate = False

            for included_entry in included:
                included_time = included_entry.get('created_at', '')

                try:
                    # Parse datetime strings for comparison
                    entry_dt = datetime.strptime(entry_time.split(' ')[0], '%Y-%m-%d')
                    included_dt = datetime.strptime(included_time.split(' ')[0], '%Y-%m-%d')

                    time_diff = abs((included_dt - entry_dt).total_seconds())

                    if time_diff < 86400:  # 24 hours in seconds
                        is_duplicate = True
                        duplicates_removed += 1
                        break
                except:
                    # If we can't parse dates, keep the entry to be safe
                    pass

            if not is_duplicate:
                included.append(entry)

        filtered_entries.extend(included)

    if duplicates_removed > 0:
        print(f"  Removed {duplicates_removed} duplicate entries within 24h (kept newest)")

    return filtered_entries


def parse_csv(input_file):
    """Parse the CSV file and organize by questionnaire type"""
    data_by_type = {
        'wlus': [],
        'wlusmonthly': [],
        'facesheet': [],
        'facesheet_ext': []
    }

    parsers = {
        'wlus': parse_wlus,
        'wlusmonthly': parse_wlusmonthly,
        'facesheet': parse_facesheet,
        'facesheet_ext': parse_facesheet_ext
    }

    skipped_types = set()
    processed_count = 0

    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_num, row in enumerate(reader, 1):
            if len(row) < 2:
                continue

            try:
                # Parse JSON data from second column
                entry = json.loads(row[1])
                qtype = entry.get('type', 'unknown')

                # Process only the types we care about
                if qtype in parsers:
                    parsed_data = parsers[qtype](entry)
                    data_by_type[qtype].append(parsed_data)
                    processed_count += 1
                else:
                    skipped_types.add(qtype)

            except json.JSONDecodeError as e:
                print(f"Warning: Could not parse JSON in row {row_num}: {e}")
            except Exception as e:
                print(f"Warning: Error processing row {row_num}: {e}")

    print(f"\nProcessed {processed_count} questionnaire entries")
    print(f"Skipped questionnaire types: {', '.join(sorted(skipped_types))}")

    # Filter duplicates within 24h for each type
    print("\nFiltering duplicates within 24h...")
    for qtype in data_by_type:
        before_count = len(data_by_type[qtype])
        data_by_type[qtype] = filter_duplicates_within_24h(data_by_type[qtype])
        after_count = len(data_by_type[qtype])
        print(f"  {qtype}: {before_count} -> {after_count} entries")

    return data_by_type


def create_excel(data_by_type, output_file):
    """Create Excel file with separate tabs for each questionnaire type"""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for qtype, entries in data_by_type.items():
            if entries:
                # Convert to DataFrame
                df = pd.DataFrame(entries)

                # Write to Excel
                df.to_excel(writer, sheet_name=qtype, index=False)

                print(f"Created tab '{qtype}' with {len(df)} rows and {len(df.columns)} columns")

    # Format headers
    wb = load_workbook(output_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Make header row bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb.save(output_file)
    print(f"\nExcel file created: {output_file}")


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 questionnaire_parser.py <input_csv>")
        sys.exit(1)

    input_file = sys.argv[1]

    if not Path(input_file).exists():
        print(f"Error: Input file '{input_file}' not found")
        sys.exit(1)

    # Create output filename
    input_path = Path(input_file)
    output_file = input_path.parent / f"{input_path.stem}_parsed.xlsx"

    print(f"Parsing questionnaire data from: {input_file}")
    print(f"Output will be saved to: {output_file}")

    # Parse CSV
    data_by_type = parse_csv(input_file)

    # Create Excel file
    create_excel(data_by_type, output_file)

    print("\nParsing complete!")


if __name__ == '__main__':
    main()
